import io
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from odoo import http
from odoo.http import request


class Binary(http.Controller):
    """Generador del llibre IGI (Andorra) en formato XLSX."""

    # Estilos base
    UNDERLINED_BOLD_FONT = Font(underline='single', bold=True)
    UNDERLINED_BOLD_FONT_RED = Font(underline='single', bold=True, color='FF0000')
    ALIGN_CENTER = Alignment(horizontal='center')
    BOLD_FONT = Font(bold=True)

    @http.route('/web/binary/xlsx_vat_book/<int:book_id>', type='http', auth='user', csrf=False)
    def xlsx_vat_book(self, book_id):
        """
        Genera y descarga el libro IGI (IVA andorrano) en formato XLSX
        para el registro l10n.ad.vat.book indicado.
        Compatible con Odoo 18.
        """
        # Recuperar el asistente (TransientModel)
        vat_book = request.env['l10n.ad.vat.book'].sudo().browse(book_id)
        if not vat_book.exists():
            return request.not_found()

        # Obtener facturas según el rango y tipo
        invoices = vat_book.get_invoices(
            vat_book.date_from, vat_book.date_to, vat_book.invoice_type
        )

        # Preparar líneas para el Excel
        invoice_lines = self._prepare_invoice_lines(invoices)

        company = request.env.user.company_id

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Llibre IGI"

        # Configurar columnas
        for col, width in zip("ABCDEFGH", [10, 15, 13, 25, 18, 15, 10, 12]):
            ws.column_dimensions[col].width = width

        # Encabezado principal
        row = 1
        ws[f"A{row}"] = company.name
        ws[f"A{row}"].font = self.UNDERLINED_BOLD_FONT_RED

        ws[f"C{row}"] = "Imprès el:"
        ws[f"C{row}"].font = self.BOLD_FONT
        ws[f"D{row}"] = time.strftime("%d/%m/%Y")
        ws[f"D{row}"].font = self.BOLD_FONT

        # Título
        row = 2
        title = "Factures EMESES" if vat_book.invoice_type == 'issued' else "Factures REBUDES"
        ws[f"E{row}"] = title
        ws[f"E{row}"].font = self.BOLD_FONT
        ws[f"E{row}"].alignment = self.ALIGN_CENTER

        # Fechas
        row = 3
        ws[f"A{row}"] = "Desde:"
        ws[f"A{row}"].font = self.UNDERLINED_BOLD_FONT
        ws[f"B{row}"] = str(vat_book.date_from)
        ws[f"B{row}"].font = self.BOLD_FONT

        row = 4
        ws[f"A{row}"] = "Fins a:"
        ws[f"A{row}"].font = self.UNDERLINED_BOLD_FONT
        ws[f"B{row}"] = str(vat_book.date_to)
        ws[f"B{row}"].font = self.BOLD_FONT

        # Cabeceras de tabla
        row = 6
        headers = [
            "Data", "Núm. Factura", "N.R.T.", "Client",
            "Import Factura", "Base Imposable", "% IGI", "Quota IGI"
        ]
        for col, title in zip("ABCDEFGH", headers):
            ws[f"{col}{row}"] = title
            ws[f"{col}{row}"].font = self.UNDERLINED_BOLD_FONT

        # Contenido
        row = 7
        added_invoices = set()
        total_invoice = total_base = total_tax = 0.0

        for line in invoice_lines:
            inv_id = line["invoice_id"]

            if inv_id not in added_invoices:
                added_invoices.add(inv_id)
                ws[f"A{row}"] = line["invoice_date"].strftime("%d/%m/%Y") if line["invoice_date"] else ""
                ws[f"B{row}"] = line["num_fac"]
                ws[f"C{row}"] = line["vat_number"]
                ws[f"D{row}"] = line["partner_name"]

            ws[f"E{row}"] = line["total_amount"]
            ws[f"F{row}"] = line["base"]
            ws[f"G{row}"] = f"{line['tax_rate']}%"
            ws[f"H{row}"] = line['amount']

            total_invoice += line["total_amount"]
            total_base += line["base"]
            total_tax += line["amount"]

            row += 1

        # Totales
        ws[f"E{row}"] = ws[f"F{row}"] = ws[f"H{row}"] = "TOTAL"
        for col in ("E", "F", "H"):
            ws[f"{col}{row}"].font = self.UNDERLINED_BOLD_FONT_RED
            ws[f"{col}{row}"].alignment = self.ALIGN_CENTER

        row += 1
        ws[f"E{row}"] = total_invoice
        ws[f"F{row}"] = total_base
        ws[f"H{row}"] = total_tax

        # Guardar en memoria
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)

        filename = f"llibre_igi_{vat_book.invoice_type}_{time.strftime('%Y%m%d')}.xlsx"
        return request.make_response(
            fp.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename={filename}')
            ]
        )

    # ---------------------------------------------------------------------
    # MÉTODO AUXILIAR
    # ---------------------------------------------------------------------
    def _prepare_invoice_lines(self, invoices):
        """Prepara las líneas de impuestos de las facturas para exportar a Excel."""
        lines = []
        for inv in invoices:
            for tax_line in inv.line_ids.filtered(lambda l: l.tax_line_id):
                # Calcular importe del impuesto
                tax_amount = abs(tax_line.balance)
                base_amount = abs(tax_line.tax_base_amount)
                total_amount = base_amount + tax_amount

                # Detectar porcentaje (directo o de impuestos hijos)
                tax_rate = tax_line.tax_line_id.amount or sum(
                    t.amount for t in tax_line.tax_line_id.children_tax_ids
                )

                lines.append({
                    'invoice_id': inv.id,
                    'invoice_date': inv.invoice_date,
                    'num_fac': inv.name,
                    'vat_number': inv.partner_id.vat or '',
                    'partner_name': inv.partner_id.name or '',
                    'base': base_amount,
                    'amount': tax_amount,            # 💰 cuota IGI (importe)
                    'tax_rate': tax_rate,            # 📊 porcentaje
                    'total_amount': total_amount,
                    'tax_name': tax_line.tax_line_id.name or '',
                })
        return lines

